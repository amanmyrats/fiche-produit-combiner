from pathlib import Path
import datetime
from fiche_produit_extracter_class import FicheProduit_to_Excel
from xls2xlsx import XLS2XLSX
import openpyxl
from openpyxl import load_workbook, workbook, Workbook
from openpyxl.styles import PatternFill, borders, Alignment
from openpyxl.styles.borders import Border, Side
import static_info
from openpyxl.utils import get_column_letter
import pandas as pd
path = Path(Path.cwd() / 'mekan/jirik/bet/gyt')

if path.exists():
    print('Bar eken')
else:
    path.mkdir(parents=True)

for xfile in path.rglob('*'):
    if xfile.is_file():
        print("File eken")
    else:
        print("papka bar eken")

# xxx=Path.cwd / 'mekan'
print(path)