import win32com.client       # Need pywin32 from pip
from PIL import ImageGrab    # Need PIL as well
import os
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import string

letters=string.ascii_uppercase


from pathlib import Path
import general_functions



start_time=datetime.datetime.now()

def image_extracter(xpath):
    dest_folder=Path(xpath)
    for excel_file in dest_folder.rglob("*.xls*"):
        try:
            workbook =load_workbook(filename=excel_file)
        except:
            continue

        for sheetname in workbook.sheetnames:
            sheet=workbook[sheetname]
            print ("Extracting images from ", excel_file.name , " - ", sheet)
            
            try:
                images._images.clear()
            except:
                pass
            finally:
                images=SheetImageLoader(sheet)

            try:
                for i in range(10,125):
                    for j in letters:
                        xcell=j + str(i)
                        if xcell in images._images.keys():
                            image=images.get(xcell)
                            image_name=str(general_functions.convert_to_file_folder_name(sheet.cell(row=2, column=7).value))+ str(xcell)+".png"
                            image.save(Path.cwd() / "cra_list_test/result_image" / image_name)
            except:
                pass

        workbook.close()

image_extracter("D:\\bckup Aman\\coding\\fiche_produit\\cra_list_test")

end_time=datetime.datetime.now()

print("Total time: ", end_time-start_time)