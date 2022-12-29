import win32com.client       # Need pywin32 from pip
from PIL import ImageGrab    # Need PIL as well
import pandas as pd 
import openpyxl
import os
import static_info
import general_functions
import datetime

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from pathlib import Path
from fp_formatter import FP_Formatter
from image_extracter_with_openpyxl_class import FP_Image_Extractor

class FicheProduit_to_Excel:
    def __init__(self, *args, **kwargs):
        # Expected arguments
        # xpath
        # result_path
        self.start_time=datetime.datetime.now()

        self.title_label_to_search=dict(static_info.title_label_to_search)
        self.list_of_colno_to_remove_from_labels=[]

        self.result_dict_for_excel=static_info.result_dict_for_excel
        self.result_path=Path(kwargs['result_path'])

        self.wb=Workbook()
        self.ws=self.wb.active
        self.result_excel_workbook_name='Fiche Produit Table.xlsx'
        self.result_excel_sheet_name='Fiche Produit Table'

        self.result_df=pd.DataFrame()
        self.fp_folder=Path(kwargs['xpath'])

        self.title_column_no_of_folder=21
        self.title_column_no_of_file=22
        self.title_column_no_of_sheetname=23
        self.title_column_no_of_empty=25

        self.excel_app=win32com.client.Dispatch('Excel.Application')
        
    def loop_trough_excels(self):
        self.counter=0
        for excel_file in self.fp_folder.rglob("*.xls*"):
            
            self.temp_file_name=str(excel_file.name)
            if "~$" in self.temp_file_name:
                print("I skipped this file: ", excel_file.name)
                continue
            
            self.parent_folder=excel_file.parts[len(excel_file.parts)-2]
            print("I will start working this file: ", excel_file.parent, " - ", excel_file.name)
            
            #Here I decide how many sheet do I have, and I assign openpyxl as well
            try:
                self.temp_wb=pd.ExcelFile(excel_file)
                self.total_sheet=len(self.temp_wb.sheet_names)
                print(type(self.temp_wb.sheet_names))
                if excel_file.suffix=='.xlsx':
                    self.file_openpyxl=load_workbook(excel_file)
                elif excel_file.suffix=='.xls':
                    print('excel_file: ', excel_file)
                    self.xls_workbook=self.excel_app.Workbooks.Open(excel_file)
                    print('xls_workbook: ', self.xls_workbook)
                    #os.system('pause')
            except:
                continue
            
            # Loop through all sheets of current workbook, excel_file
            for i in range(0, self.total_sheet):
                # # Here I assign label dictionary at the beginning of every sheet
                # self.title_label_to_search_copy=self.title_label_to_search
                try:
                    self.df=pd.read_excel(excel_file, sheet_name=self.temp_wb.sheet_names[i])
                    print("Dataframe is assigned to sheet", self.temp_wb.sheet_names[i])
                except:
                    print('Sheet could not be assigned to dataframe: ', self.temp_wb.sheet_names[i])
                    continue

                # Decide if the sheet is fiche produit or not
                if not general_functions.is_sheet_fiche_produit(type='dataframe', df=self.df, excel_file=excel_file, sheet_name=self.temp_wb.sheet_names[i]):
                    print("This sheet --- {} is not a fiche produit sheet".format(self.temp_wb.sheet_names[i]))

                    # Here I want to track excel files which are skipped.
                    # Independent from loop, update folder name, excel file name and current sheet name for once here
                    self.counter+=1
                    self.result_dict_for_excel[self.title_column_no_of_folder].append(self.parent_folder)
                    self.result_dict_for_excel[self.title_column_no_of_file].append(excel_file.stem)
                    self.result_dict_for_excel[self.title_column_no_of_sheetname].append(self.temp_wb.sheet_names[i])
                    self.result_dict_for_excel[self.title_column_no_of_empty].append('Empty')

                    # To make all columns same size, I write "" into every blank cell here
                    self.make_all_same_size()

                    continue
                else:
                    # This is a fiche produit sheet, so increase counter by 1
                    self.counter+=1

                    # It is test here
                    self.title_label_to_search=dict(static_info.title_label_to_search)

                    # For each row in dataframe
                    for k, rows in self.df.iterrows():
                        # For each cell in row
                        for j, cols in rows.iteritems():
                            self.column_no=self.df.columns.get_loc(j)
                            self.xvalue=str(cols)
                            
                            if str(self.xvalue)=="nan":
                                continue

                            if self.column_no>25:
                                continue

                            # Remove keys those have been found already
                            try:
                                self.dict_copy=dict(self.title_label_to_search)
                                for to_remove in self.list_of_colno_to_remove_from_labels:
                                    for key, value in self.title_label_to_search.items():
                                        if value[1]==to_remove:
                                            self.dict_copy.pop(key)
                                self.title_label_to_search=dict(self.dict_copy)
                            except:
                                pass

                            self.list_of_colno_to_remove_from_labels=[]
                            # Check if cell value is equal to column titles
                            for key in self.title_label_to_search:

                                if self.xvalue.__contains__(key) and len(self.xvalue)<80:
                                    self.column_no_in_result_dict=int(self.title_label_to_search[key][1])
                                    if self.title_label_to_search[key][0]=="regular":
                                        self.current_cell_value=self.df.iloc[k][self.column_no]
                                        
                                        # Call regular search function and assign value
                                        self.regular_search(k=k)
                                        if self.desired_value_found:
                                            self.list_of_colno_to_remove_from_labels.append(self.column_no_in_result_dict)
                                            break
                                        
                                    elif self.title_label_to_search[key][0]=="not_regular": # if it is not regular, it is for description of product.
                                        try:
                                            self.not_regular_search(k=k)
                                            # Keys to remove, because they were found already
                                            self.list_of_colno_to_remove_from_labels.append(self.column_no_in_result_dict)
                                        except:
                                            print("Some error occured. When workig with techicanl description.")
                                            pass

                                    elif self.title_label_to_search[key][0]=="footer":     # if it is footer
                                        # Check the second value of footer
                                        self.concataneted_footer=""
                                        if self.xvalue.__contains__(self.title_label_to_search[key][2]):
                                            self.max_length=self.title_label_to_search[key][3]
                                            try:
                                                self.footer_search(k=k)
                                                # Keys to remove, because they were found already
                                                self.list_of_colno_to_remove_from_labels.append(self.column_no_in_result_dict)
                                            except:
                                                print("Some error occured. When working with footer.")
                                                pass

                # Independent from loop, update folder name, excel file name and current sheet name for once here
                self.result_dict_for_excel[self.title_column_no_of_folder].append(self.parent_folder)
                self.result_dict_for_excel[self.title_column_no_of_file].append(excel_file.stem)
                self.result_dict_for_excel[self.title_column_no_of_sheetname].append(self.temp_wb.sheet_names[i])

                # To make all columns same size, I write "" into every blank cell here
                self.make_all_same_size()
                
                # Extract images
                self.extract_workbook_images(i_loop=i, excel_file=excel_file)
                
            # Close openpyxl excel file after done
            if excel_file.suffix=='.xlsx':
                self.file_openpyxl.close()
            elif excel_file.suffix=='.xls' and not self.xls_workbook==None:
                self.xls_workbook.Close(False)
             # Set pywin32 excel app to none
            try:
                self.excel_app.Quit
            except:
                print('Error when trying to Quit win32com object')
            finally:
                self.excel_app=None

        print("Everything is finished, now I will assign result_df")
        print("Length of result_dicts are:")
        for i in range(1, len(self.result_dict_for_excel)):
            print(i, " - ", len(self.result_dict_for_excel[i]))

        # Here I create empty dataframe and assign result dictionary into in it
        self.result_df=pd.DataFrame()
        for i in range(0, len(self.result_dict_for_excel)):
            self.result_df[static_info.formal_title_list[i]]=self.result_dict_for_excel[i+1]

        #Save dataframe into excel file
        self.result_df.to_excel(self.result_path / self.result_excel_workbook_name)

        # Format result excel table
        self.wb_to_format=load_workbook(self.result_path / self.result_excel_workbook_name)
        self.wb_to_format['Sheet1'].title=self.result_excel_sheet_name
        self.formatter=FP_Formatter(wb=self.wb_to_format, save_path=self.result_path, wb_name=self.result_excel_workbook_name, sh_name=self.result_excel_sheet_name)

        # Clear the result dictionary after saving
        for key in self.result_dict_for_excel:
            self.result_dict_for_excel[key].clear()

        # Calculate total time and show.
        self.end_time=datetime.datetime.now()
        self.total_time_spent=self.end_time-self.start_time
        print("Total spent time is: ", self.total_time_spent)




    def make_all_same_size(self):
        self.total_col_must_be=len(self.result_dict_for_excel)
        for colno in range(1 , self.total_col_must_be+1):
            self.total_row_current=len(self.result_dict_for_excel[colno])
            if self.total_row_current>self.counter:
                print('Should not exceed the counter.')
                pass
            if self.total_row_current<self.counter:
                # If current column number is less than expected column number, then add "" to make it even
                self.result_dict_for_excel[colno].append("-")


    def extract_workbook_images(self, **kwargs):
        # Extract images
        self.image_name=self.result_dict_for_excel[3][len(self.result_dict_for_excel[3])-1]
        print('Image name: ', self.image_name)
        self.fp_numero=str(self.result_dict_for_excel[19][len(self.result_dict_for_excel[19])-1])
        self.fp_indice=str(self.result_dict_for_excel[20][len(self.result_dict_for_excel[20])-1])
        self.none_to_zero=lambda x:'0' if x=='-' else x
        self.fp_no=self.fp_numero + '-' + self.none_to_zero(self.fp_indice)
        self.excel_name=Path(kwargs['excel_file']).name
        #print('Suffix: ', Path(kwargs['excel_file']).suffix)
        if Path(kwargs['excel_file']).suffix=='.xlsx':
            obj_extractor=FP_Image_Extractor(xpath=self.fp_folder, result_path=self.result_path)
            obj_extractor.xlsx_image_extracter(workbook=self.file_openpyxl, sheetname=self.temp_wb.sheet_names[kwargs['i_loop']], image_name=self.image_name, fp_no=self.fp_no, excel_name=self.excel_name, df_for_excel=self.result_dict_for_excel)
            obj_extractor=None
        elif Path(kwargs['excel_file']).suffix=='.xls' and not self.xls_workbook==None:
            obj_extractor=FP_Image_Extractor(xpath=self.fp_folder, result_path=self.result_path)
            obj_extractor.xls_image_extracter(workbook=self.xls_workbook, sheetname=self.xls_workbook.Worksheets(self.temp_wb.sheet_names[kwargs['i_loop']]), image_name=self.image_name, fp_no=self.fp_no, excel_name=self.excel_name, df_for_excel=self.result_dict_for_excel)
            obj_extractor=None

        # self.image_name=""
        # self.fp_numero=""
        # self.fp_indice=""
        # self.fp_no=""

    def regular_search(self, **kwargs):
        self.desired_value_found=False
        for to_right in range(self.column_no+1, 25):
            try:
                self.currently_active_iloc_value=self.df.iloc[kwargs['k']][to_right]
                if general_functions.is_there_value(self.currently_active_iloc_value):
                    self.result_dict_for_excel[self.column_no_in_result_dict].append(self.currently_active_iloc_value)
                    self.desired_value_found=True
                    break
            except:
                pass
            

    def not_regular_search(self, **kwargs):
        self.description_found_counter=0
        self.full_description=''
        for to_right in range(self.column_no, 25):
            self.currently_active_iloc_value=str(self.df.iloc[kwargs['k']+1][to_right])
            self.current_cell_value=str(self.df.iloc[kwargs['k']][self.column_no])
            self.description_found_counter+=1
            if self.description_found_counter>5:
                self.result_dict_for_excel[self.column_no_in_result_dict].append(self.full_description)
                break
            elif not str(self.currently_active_iloc_value).lower()=='nan':
                self.full_description+=self.currently_active_iloc_value


    def footer_search(self, **kwargs):
        for to_right in range(self.column_no, 25):
            self.currently_active_iloc_value=str(self.df.iloc[kwargs['k']+1][to_right])
            self.current_cell_value=str(self.df.iloc[kwargs['k']][self.column_no])
            if general_functions.is_there_value(self.currently_active_iloc_value):
                if len(self.concataneted_footer)<self.max_length:
                    if len(self.currently_active_iloc_value)>0:
                        if str(self.currently_active_iloc_value).__contains__("."):
                            self.dot_position=str(self.currently_active_iloc_value).find(".")
                            self.concataneted_footer+=str(self.currently_active_iloc_value)[0:self.dot_position]
                        else:
                            self.concataneted_footer+=str(self.currently_active_iloc_value)
                else:
                    self.result_dict_for_excel[self.column_no_in_result_dict].append(self.concataneted_footer)
                    break