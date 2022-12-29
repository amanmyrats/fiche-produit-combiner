from pathlib import Path
import openpyxl
from openpyxl import load_workbook, workbook
from openpyxl.styles import PatternFill, borders, Alignment
from openpyxl.styles.borders import Border, Side
import static_info
from openpyxl.utils import get_column_letter

class FP_Formatter:
    def __init__(self, *args, **kwargs):
        self.wb=kwargs['wb']
        self.wb_name=kwargs['wb_name']
        self.sh_name=kwargs['sh_name']

        self.save_path=kwargs['save_path']
        self.sh=self.wb[self.sh_name]
        # Auto-Filter all
        self.sh.auto_filter.ref=self.sh.dimensions

        # Set width of all columns according to title label
        self.title_label_dict=dict(static_info.title_label_to_search)
        for key, value in self.title_label_dict.items():
            self.sh.column_dimensions[get_column_letter(value[1]+1)].width=value[4]

        # Border styles
        self.dotted_border=Border(left=Side(style='dotted'), right=Side('dotted'), top=Side('dotted'), bottom=Side('dotted'))
        self.thin_border=Border(left=Side(style='thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        # Wrap the names, column D and E.
        for col_to_wrap in self.sh['D':'E']:
            for cell_to_wrap in col_to_wrap:
                #print('cell to wrap: ', cell_to_wrap)
                cell_to_wrap.alignment=Alignment(wrap_text=True)

        # Change style of all cells
        for rows in self.sh[self.sh.dimensions]:
            for xcell in rows:
                xcell.border=self.dotted_border
                xcell.alignment=Alignment(vertical='center')

        # Columns needs to be centered
        self.columns_to_center=('f', 'n', 'o', 'p', 'q', 'r', 's')
        for col_to_center in self.columns_to_center:
            for cell_to_center in self.sh[col_to_center]:
                cell_to_center.alignment=Alignment(horizontal='center')

        # Change style of columns which contains file/folder information
        self.columns_not_in_dict=('u', 'v', 'w', 'x', 'y')
        for col_not_in_dict in self.columns_not_in_dict:
            self.sh.column_dimensions[col_not_in_dict].width=20


        # Style changes of title row
        self.column_width_set=False
        self.sh.row_dimensions[1].height=50
        for title_cell in self.sh[self.sh.dimensions][0]:
            # Color
            title_cell.fill=PatternFill(start_color='FF87CEFA', end_color='FF87CEFA', fill_type='solid')
            # Border
            title_cell.border=self.thin_border
            # Alignment and Wrap
            title_cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)

        self.wb.save(self.save_path / self.wb_name)
        self.wb.close()
        print('Finished formatting.')

